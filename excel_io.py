"""
Excel 读取与美化输出
"""
import os
from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import AssetRow, JudgeResult, AssetGroup, WorkflowState
from .config import OUTPUT_DIR


# ── 样式常量 ──
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="微软雅黑", size=10)
DATA_ALIGNMENT = Alignment(vertical="center", wrap_text=True)

YES_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NO_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
REVIEW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# 交替行颜色
ALT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")


def _style_header(ws, num_cols: int):
    """应用表头样式"""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _style_data(ws, num_rows: int, num_cols: int):
    """应用数据区域样式"""
    for row in range(2, num_rows + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER
            if row % 2 == 0:
                cell.fill = ALT_FILL


def _auto_width(ws, num_cols: int, max_width: int = 40):
    """自动列宽"""
    for col in range(1, num_cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                val = str(cell.value or "")
                # 中文字符占两个宽度
                char_len = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, char_len)
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, max_width)


def _freeze_and_filter(ws):
    """冻结首行 + 自动筛选"""
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ────────────────────────────────────────
# 读取输入 Excel
# ────────────────────────────────────────
def read_input_excel(filepath: str) -> list[AssetRow]:
    """读取输入Excel，返回 AssetRow 列表"""
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is None:
            continue
        rows.append(AssetRow(
            row_index=idx,
            topic=str(row[0] or "").strip(),
            category_l2=str(row[1] or "").strip(),
            category_l3=str(row[2] or "").strip(),
            asset=str(row[3] or "").strip(),
        ))
    wb.close()
    print(f"[输入] 读取 {len(rows)} 行数据")
    return rows


# ────────────────────────────────────────
# 输出 01: 同名聚合
# ────────────────────────────────────────
def write_phase0_excel(state: WorkflowState):
    """输出阶段0结果：精确同名分组"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "精确同名聚合"

    headers = ["组号", "归一化名称", "原始资产名", "出现次数", "所属主题域"]
    ws.append(headers)

    groups = state["normalized_groups"]
    group_id = 0
    for norm_name, rows in sorted(groups.items()):
        distinct_names = sorted(set(r.asset for r in rows))
        if len(rows) < 2:
            continue
        group_id += 1
        topics = sorted(set(r.topic for r in rows))
        for name in distinct_names:
            count = sum(1 for r in rows if r.asset == name)
            ws.append([group_id, norm_name, name, count, " / ".join(topics)])

    num_rows = ws.max_row
    num_cols = len(headers)
    _style_header(ws, num_cols)
    _style_data(ws, num_rows, num_cols)
    _auto_width(ws, num_cols)
    _freeze_and_filter(ws)

    path = os.path.join(OUTPUT_DIR, "01_同名聚合.xlsx")
    wb.save(path)
    print(f"[输出] {path}")


# ────────────────────────────────────────
# 输出 02: 候选对 + AI 判断结果
# ────────────────────────────────────────
def write_phase23_excel(state: WorkflowState):
    """输出阶段2+3结果：候选对 + 判断结果"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "候选对判断结果"

    headers = ["资产A", "资产B", "判断结果", "来源"]
    ws.append(headers)

    edges: list[JudgeResult] = state["edges"]
    for e in sorted(edges, key=lambda x: (x.result, x.name_a)):
        ws.append([e.name_a, e.name_b, e.result, e.source])

    num_rows = ws.max_row
    num_cols = len(headers)
    _style_header(ws, num_cols)
    _style_data(ws, num_rows, num_cols)

    # 条件格式：YES 绿色，NO 红色
    for row in range(2, num_rows + 1):
        cell = ws.cell(row=row, column=3)
        if cell.value == "YES":
            cell.fill = YES_FILL
        elif cell.value == "NO":
            cell.fill = NO_FILL

    _auto_width(ws, num_cols)
    _freeze_and_filter(ws)

    path = os.path.join(OUTPUT_DIR, "02_候选对判断.xlsx")
    wb.save(path)
    print(f"[输出] {path}")


# ────────────────────────────────────────
# 输出 03: 相似资产图（最终结果）
# ────────────────────────────────────────
def write_final_excel(state: WorkflowState):
    """输出最终结果：边表 + 组表 + 明细"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = Workbook()

    # ── Sheet1: 边表 ──
    ws1 = wb.active
    ws1.title = "边表"
    headers1 = ["资产A", "资产B", "判断结果", "来源"]
    ws1.append(headers1)

    edges: list[JudgeResult] = state["edges"]
    yes_edges = [e for e in edges if e.result == "YES"]
    for e in sorted(yes_edges, key=lambda x: x.name_a):
        ws1.append([e.name_a, e.name_b, e.result, e.source])

    _style_header(ws1, len(headers1))
    _style_data(ws1, ws1.max_row, len(headers1))
    for row in range(2, ws1.max_row + 1):
        ws1.cell(row=row, column=3).fill = YES_FILL
    _auto_width(ws1, len(headers1))
    _freeze_and_filter(ws1)

    # ── Sheet2: 组表 ──
    ws2 = wb.create_sheet("组表")
    headers2 = ["组号", "组大小", "资产名列表", "是否完全连通", "是否需人工审核"]
    ws2.append(headers2)

    groups: list[AssetGroup] = state["groups"]
    for g in sorted(groups, key=lambda x: -len(x.asset_names)):
        ws2.append([
            g.group_id,
            len(g.asset_names),
            " | ".join(g.asset_names),
            "是" if g.is_fully_connected else "否",
            "是" if g.needs_review else "否",
        ])

    _style_header(ws2, len(headers2))
    _style_data(ws2, ws2.max_row, len(headers2))
    # 标记需审核的行
    for row in range(2, ws2.max_row + 1):
        if ws2.cell(row=row, column=5).value == "是":
            for col in range(1, len(headers2) + 1):
                ws2.cell(row=row, column=col).fill = REVIEW_FILL
    _auto_width(ws2, len(headers2))
    _freeze_and_filter(ws2)

    # ── Sheet3: 明细（映射回原始行）──
    ws3 = wb.create_sheet("明细")
    headers3 = ["组号", "资产名", "主题域", "二级分类", "三级分类", "数据资产(原始)"]
    ws3.append(headers3)

    # 建立资产名 → 组号映射
    name_to_group = {}
    for g in groups:
        for name in g.asset_names:
            name_to_group[name] = g.group_id

    # 映射回原始行：
    # 1) 已分组的全部保留
    # 2) 未分组但在输入中重复出现(>1次)的也保留
    asset_rows: list = state["asset_rows"]
    name_counts = defaultdict(int)
    for row in asset_rows:
        name_counts[row.asset] += 1

    detail_rows = []
    for row in asset_rows:
        gid = name_to_group.get(row.asset)
        if gid is not None or name_counts[row.asset] > 1:
            detail_rows.append((gid, row))

    def _detail_sort_key(item):
        gid, row = item
        gid_sort = gid if gid is not None else 10**9
        return gid_sort, row.asset

    for gid, row in sorted(detail_rows, key=_detail_sort_key):
        gid_value = gid if gid is not None else "重复未成组"
        ws3.append([gid_value, row.asset, row.topic, row.category_l2, row.category_l3, row.asset])

    _style_header(ws3, len(headers3))
    _style_data(ws3, ws3.max_row, len(headers3))
    _auto_width(ws3, len(headers3))
    _freeze_and_filter(ws3)

    path = os.path.join(OUTPUT_DIR, "03_相似资产图.xlsx")
    wb.save(path)
    print(f"[输出] {path}")
